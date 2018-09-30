import sys
import random
import argparse
import random
from openpyxl import load_workbook


def random_sized_chunks(n_chunks, length):
    """
    Generate random stop points for breaking word into chunks.
    This function will generate the `end` part of the start-end range
    """

    stop_points = []
    for i in range(n_chunks - 1):
        if len(stop_points):
            start = stop_points[-1]
        else:
            start = 0
        point = random.randint(start, length - n_chunks + i)
        while point == start:
            # dont want to have empty chunks
            point = random.randint(start, length - n_chunks + i)
        stop_points.append(point)
    # last stop point is the end of the word
    stop_points.append(length)
    return stop_points


def random_sized_subwords(n_subwords, word):
    """
    Using a list of start-end points, generate a list of size n_subwords, for a given word
    """
    chunks = random_sized_chunks(n_subwords, len(word))
    start_end_points = zip([0] + chunks[:len(chunks) - 1], chunks)
    for start, end in start_end_points:
        yield word[start:end]


def split_half(word):
    middle = len(word) // 2
    return [word[: middle],  word[middle:]]


def two_letter_subwords(word):
    for i in range(0, len(word), 2):
        yield word[i:i+2]


def generate_word_subwords(word):
    # "Two sub words where the sub words are divided into equally long words"
    two_subs = split_half(word)
    third_length = len(word) // 3
    # "Three sub words where the sub words are divided into equally long words"
    three_subs = [word[: third_length], word[third_length: third_length * 2], word[third_length *2:]]
    # "The sub words are divided into single characters each, if the word is maximum 6 chars long"
    if len(word) <= 6:
        single_chars = [ch for ch in word]
    else:
        single_chars = []
    # The sub words are divided into two characters each
    two_letter_subs = list(two_letter_subwords(word))

    random_subs = []
    # "For words longer than 7 characters: The sub words are divided into 4 random length characters."
    # Repeat for 9-5, 11-6, 13-7
    random_sub_key = 0
    random_subword_guide = [(13, 7, 8), (11, 6, 7), (9, 5, 6), (7, 4, 5)]
    for min_len, n_subs, key in random_subword_guide:
        if len(word) > min_len:
            random_subs = list(random_sized_subwords(n_subs, word))
            random_sub_key = key
            # This rule should be applied only once
            break

    return {
        1: two_subs,
        2: three_subs,
        3: single_chars,
        4: two_letter_subs,
        random_sub_key: random_subs
    }


def format_result(type_of_subwords, subwords_list, category, word):
    if not type_of_subwords or not subwords_list:
        return None
    subs = '/'.join(subwords_list)
    out = {
        'category': category,
        'type_of_subwords': type_of_subwords,
        'word': word,
        'num_of_subwords': len(subwords_list),
        'subs': subs
    }
    return '{category}/{type_of_subwords}/{num_of_subwords}/{word}/{subs}'.format(**out)


def process_sheet(inputfilename, sheetname):
    #loading the inputfile
    workbook = load_workbook(filename=inputfilename, read_only=True)

    #check the sheetname parameter
    if sheetname is not None:
        sheet = workbook.get_sheet_by_name(name=sheetname)
    else:
        sheet = workbook.active

    #list for the results
    results = list()
    #iterating through the rows expect the first row(it contains the category names)
    for row in sheet.iter_rows(row_offset=1):
        #iterating through each cell in the row
        for cell in row:
            if cell.value is None:
                # skip empty cells
                continue
            #calculate category
            category =  str(cell.column)
            word = cell.value

            all_subwords = generate_word_subwords(word)
            for type_of_subwords, subwords in all_subwords.items():
                formatted_result = format_result(type_of_subwords, subwords, category, word)
                if formatted_result:
                    results.append(formatted_result)

    return results


def save_results(results, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        #sort by category
        results = sorted(results, key=lambda word_line: word_line[0])
        #save the results
        for line in results:
            print(line, file=f)
    print('Creating subwords has finished. Saved result to {}'.format(filename))


parser = argparse.ArgumentParser(description='Generate subwords from input XLS sheet')
parser.add_argument('input', help='The XLSX file containing all the words')
parser.add_argument('--sheet', default=None, help='The name of the sheet within input file')
parser.add_argument('--out', default='results.txt', help='The name of the file where results are saved')

if __name__ == '__main__':

    args = parser.parse_args()
    results = process_sheet(args.input, args.sheet)
    save_results(results, args.out)
